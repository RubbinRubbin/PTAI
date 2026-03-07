from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
from io import BytesIO

def export_athlete_to_excel(athlete):
    wb = Workbook()

    # Stili
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11, color="4472C4")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== FOGLIO 1: PROFILO ATLETA ==========
    ws_profile = wb.active
    ws_profile.title = "Profilo Atleta"

    ws_profile['A1'] = f"Profilo: {athlete.nome} {athlete.cognome}"
    ws_profile['A1'].font = title_font
    ws_profile.merge_cells('A1:B1')

    profile_data = [
        ("Nome", athlete.nome),
        ("Cognome", athlete.cognome),
        ("Età", athlete.eta),
        ("Sport", athlete.sport),
        ("Altezza (cm)", athlete.altezza or "N/D"),
        ("Peso (kg)", athlete.peso or "N/D"),
        ("Data di nascita", athlete.data_nascita.strftime("%d/%m/%Y") if athlete.data_nascita else "N/D"),
        ("Descrizione", athlete.descrizione or "N/D"),
        ("Note", athlete.note or "N/D"),
    ]

    for i, (label, value) in enumerate(profile_data, start=3):
        ws_profile[f'A{i}'] = label
        ws_profile[f'A{i}'].font = header_font
        ws_profile[f'B{i}'] = value

    ws_profile.column_dimensions['A'].width = 20
    ws_profile.column_dimensions['B'].width = 40

    # ========== FOGLIO 2: RIEPILOGO METRICHE ==========
    ws_summary = wb.create_sheet("Riepilogo Metriche")

    summary_headers = ["Metrica", "Note", "Asse X", "Unità X", "Asse Y", "Unità Y", "N. Valori"]
    for col, header in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for row_idx, definition in enumerate(athlete.metric_definitions, start=2):
        ws_summary.cell(row=row_idx, column=1, value=definition.nome)
        ws_summary.cell(row=row_idx, column=2, value=definition.note or "")
        ws_summary.cell(row=row_idx, column=3, value=definition.asse_x_nome)
        ws_summary.cell(row=row_idx, column=4, value=definition.asse_x_unita)
        ws_summary.cell(row=row_idx, column=5, value=definition.asse_y_nome)
        ws_summary.cell(row=row_idx, column=6, value=definition.asse_y_unita)
        ws_summary.cell(row=row_idx, column=7, value=len(definition.values))

        for col in range(1, 8):
            ws_summary.cell(row=row_idx, column=col).border = thin_border

    for col in range(1, 8):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18

    # ========== FOGLI 3+: UN FOGLIO PER OGNI METRICA CON GRAFICO ==========
    chart_colors = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47", "264478", "9B59B6"]

    for def_idx, definition in enumerate(athlete.metric_definitions):
        # Crea nome foglio sicuro (max 31 chars, no caratteri speciali)
        sheet_name = definition.nome[:28].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '').replace(':', '-')
        # Evita nomi duplicati
        existing_names = [ws.title for ws in wb.worksheets]
        if sheet_name in existing_names:
            sheet_name = f"{sheet_name[:25]}_{def_idx}"

        ws = wb.create_sheet(sheet_name)

        # Titolo metrica
        ws['A1'] = definition.nome
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        # Note metrica
        if definition.note:
            ws['A2'] = definition.note
            ws['A2'].font = Font(italic=True, color="666666")
            ws.merge_cells('A2:D2')
            data_start_row = 4
        else:
            data_start_row = 3

        # Info assi
        ws.cell(row=data_start_row - 1, column=1, value=f"Asse X: {definition.asse_x_nome} ({definition.asse_x_unita})")
        ws.cell(row=data_start_row - 1, column=1).font = subtitle_font

        ws.cell(row=data_start_row - 1, column=3, value=f"Asse Y: {definition.asse_y_nome} ({definition.asse_y_unita})")
        ws.cell(row=data_start_row - 1, column=3).font = subtitle_font

        # Header tabella dati
        headers = [
            f"{definition.asse_x_nome} ({definition.asse_x_unita})",
            f"{definition.asse_y_nome} ({definition.asse_y_unita})",
            "Data Inserimento",
        ]
        header_row = data_start_row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Dati valori ordinati per asse X
        sorted_values = sorted(definition.values, key=lambda v: v.valore_x)
        for val_idx, value in enumerate(sorted_values):
            row = header_row + 1 + val_idx
            ws.cell(row=row, column=1, value=value.valore_x)
            ws.cell(row=row, column=2, value=value.valore_y)
            ws.cell(row=row, column=3, value=value.created_at.strftime("%d/%m/%Y %H:%M") if value.created_at else "")

            for col in range(1, 4):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

        # Larghezza colonne
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

        # ---- GRAFICO EXCEL ----
        if len(sorted_values) >= 2:
            chart = ScatterChart()
            chart.title = definition.nome
            chart.x_axis.title = f"{definition.asse_x_nome} ({definition.asse_x_unita})"
            chart.y_axis.title = f"{definition.asse_y_nome} ({definition.asse_y_unita})"
            chart.width = 20
            chart.height = 12
            chart.style = 2

            data_end_row = header_row + len(sorted_values)

            x_values = Reference(ws, min_col=1, min_row=header_row + 1, max_row=data_end_row)
            y_values = Reference(ws, min_col=2, min_row=header_row + 1, max_row=data_end_row)

            series = Series(y_values, x_values, title=definition.nome)
            color = chart_colors[def_idx % len(chart_colors)]
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = 25000  # 2pt in EMUs
            series.marker.symbol = "circle"
            series.marker.size = 6
            series.marker.graphicalProperties.solidFill = color
            series.marker.graphicalProperties.line.solidFill = color

            chart.series.append(series)

            # Posiziona il grafico a destra della tabella
            chart_cell = f"E{data_start_row}"
            ws.add_chart(chart, chart_cell)

    # Salva in buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
